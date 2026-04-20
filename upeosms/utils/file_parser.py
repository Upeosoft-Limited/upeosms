import csv
import os
from openpyxl import load_workbook

import frappe

from upeosms.utils.template import normalize_key

REQUIRED_COLUMN = "mobile"

def read_uploaded_rows(file_url: str):
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()

    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        rows, columns = _read_csv(file_path)
    elif ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        rows, columns = _read_xlsx(file_path)
    else:
        frappe.throw("Only CSV and XLSX files are supported.")

    if REQUIRED_COLUMN not in columns:
        frappe.throw("Uploaded file must contain a 'mobile' column.")

    return rows, columns

def _read_csv(file_path):
    rows = []
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(_normalize_row(row))

    columns = list(rows[0].keys()) if rows else []
    return rows, columns

def _read_xlsx(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return [], []

    headers = [normalize_key(h) for h in raw_rows[0] if h is not None]
    rows = []

    for values in raw_rows[1:]:
        row = {}
        for i, header in enumerate(headers):
            row[header] = values[i] if i < len(values) else None
        rows.append(_normalize_row(row))

    return rows, headers

def _normalize_row(row):
    normalized = {}
    for key, value in (row or {}).items():
        nk = normalize_key(key)
        if nk:
            normalized[nk] = value
    return normalized